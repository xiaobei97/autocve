"""
CVE数据抓取与处理工具
作者：Cursor AI 协助开发
功能：自动抓取指定日期范围内的CVE漏洞数据，并生成格式化的Excel报告
"""

import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import re, json
import os
import glob
from pathlib import Path
try:
    from openai import OpenAI
except ImportError:
    OpenAI = None

def get_date_input():
    while True:
        try:
            start_date = input("请输入开始日期 (格式: YYYY-MM-DD): ")
            datetime.strptime(start_date, '%Y-%m-%d')
            end_date = input("请输入结束日期 (格式: YYYY-MM-DD): ")
            datetime.strptime(end_date, '%Y-%m-%d')
            if datetime.strptime(end_date, '%Y-%m-%d') < datetime.strptime(start_date, '%Y-%m-%d'):
                print("错误：结束日期不能早于开始日期，请重新输入")
                continue
            return start_date, end_date
        except ValueError:
            print("错误：日期格式不正确，请使用 YYYY-MM-DD 格式")
            continue

def fetch_cve_data(component, start_date, end_date):
    url = "https://cvefeed.io/api/vulnerability/advanced-search"
    params = {
        "keyword": component,
        "published_after": start_date,
        "published_before": end_date,
        "cvss_min": "7.00",
        "cvss_max": "10.00",
        "order_by": "-published"
    }
    headers = {
        "Host": "cvefeed.io",
        "Connection": "keep-alive",
        "sec-ch-ua-platform": '"Windows"',
        "X-Requested-With": "XMLHttpRequest",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36",
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "sec-ch-ua": '"Not)A;Brand";v="8", "Chromium";v="138", "Google Chrome";v="138"',
        "sec-ch-ua-mobile": "?0",
        "Sec-Fetch-Site": "same-origin",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Dest": "empty",
        "Referer": f"https://cvefeed.io/search?keyword={component}&published_after={start_date}&published_before={end_date}&cvss_min=7.00&cvss_max=10.00&order_by=-published",
        "Accept-Encoding": "gzip, deflate, br, zstd",
        "Accept-Language": "zh-CN,zh;q=0.9"
    }
    try:
        print(f"正在获取组件 {component} 的CVE数据...")
        resp = requests.get(url, params=params, headers=headers, timeout=20)
        resp.raise_for_status()
        data = json.loads(resp.content.decode('utf-8'))
        return {'component': component, 'results': data.get('results', [])}
    except Exception as e:
        print(f"组件 {component} 抓取失败: {e}")
        return {'component': component, 'results': []}

def write_excel(data):
    if not data:
        print("警告：没有找到任何数据")
        return None
    rows = []
    for d in data:
        comp = d['component']
        for cve in d['results']:
            desc = re.sub('<.*?>', '', cve['description'])
            desc = re.sub(r'\s+', ' ', desc).strip()
            rows.append({
                '组件名称': comp,
                'CVE编号': cve['id'],
                '发布时间': cve['published'].split('T')[0],
                '漏洞描述': desc,
                '漏洞级别': 'High'
            })
    if not rows:
        print("警告：没有找到任何数据")
        return None
    df = pd.DataFrame(rows)
    filename = f'cve_results_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    df.to_excel(filename, index=False)
    wb = load_workbook(filename)
    ws = wb.active
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 100
    ws.column_dimensions['E'].width = 12
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col in range(1, 6):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    data_font = Font(name='微软雅黑', size=10)
    for row in range(2, len(rows) + 2):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.border = thin_border
            if col == 4:
                cell.alignment = Alignment(wrapText=True, vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 30
    comp_col = [ws.cell(row=r, column=1).value for r in range(2, len(rows)+2)]
    start = 2
    for i in range(1, len(comp_col)):
        if comp_col[i] != comp_col[i-1]:
            if i+1-start > 1:
                ws.merge_cells(f'A{start}:A{i+1}')
                ws.cell(row=start, column=1).alignment = Alignment(horizontal='center', vertical='center')
            start = i+2
    if start <= len(rows):
        if len(rows)+1-start >= 1:
            ws.merge_cells(f'A{start}:A{len(rows)+1}')
            ws.cell(row=start, column=1).alignment = Alignment(horizontal='center', vertical='center')
    wb.save(filename)
    print(f"结果已成功保存到 {filename} 文件中")
    return filename

def read_api_key():
    key_file = Path("key.txt")
    if not key_file.exists():
        raise FileNotFoundError("key.txt文件不存在")
    with open(key_file, 'r', encoding='utf-8') as f:
        api_key = f.read().strip().lstrip('\ufeff')
    if not api_key:
        raise ValueError("API key不能为空")
    return api_key

def read_prompt():
    prompt_file = Path("prompt.txt")
    if not prompt_file.exists():
        raise FileNotFoundError("prompt.txt文件不存在")
    with open(prompt_file, 'r', encoding='utf-8') as f:
        return f.read().strip()

def upload_and_analyze_bailian(xlsx_path, prompt):
    if OpenAI is None:
        print("未安装openai库，无法上传和分析。请先 pip install openai")
        return
    try:
        api_key = read_api_key()
        client = OpenAI(
            api_key=api_key,
            base_url="https://dashscope.aliyuncs.com/compatible-mode/v1"
        )
        print(f"正在上传 {xlsx_path} 到百炼平台...")
        with open(xlsx_path, 'rb') as f:
            resp = client.files.create(file=f, purpose='file-extract')
            file_id = resp.id
        print(f"上传成功，文件ID: {file_id}")
        messages = [
            {'role': 'system', 'content': 'You are a helpful assistant.'},
            {'role': 'system', 'content': f'fileid://{file_id}'},
            {'role': 'user', 'content': prompt}
        ]
        print("正在调用大模型分析，请稍候...")
        completion = client.chat.completions.create(
            model="qwen-long",
            messages=messages,
            stream=True,
            stream_options={"include_usage": True}
        )
        result = ""
        for chunk in completion:
            if chunk.choices and chunk.choices[0].delta.content:
                content = chunk.choices[0].delta.content
                print(content, end='', flush=True)
                result += content
        print("\n分析完成。结果已输出。")
        with open("bailian_analysis.txt", "w", encoding="utf-8") as f:
            f.write(result)
        print("分析结果已保存到 bailian_analysis.txt")
        # 删除百炼平台所有已上传文件
        try:
            print("正在清理百炼平台所有已上传文件...")
            response = client.files.list()
            files = response.data
            if not files:
                print("百炼平台无历史文件。")
            else:
                success_count = 0
                error_count = 0
                for file in files:
                    try:
                        client.files.delete(file.id)
                        print(f"平台文件删除成功：{file.id}")
                        success_count += 1
                    except Exception as e:
                        print(f"平台文件删除失败 {file.id}: {e}")
                        error_count += 1
                print(f"平台文件清理完成，成功：{success_count}，失败：{error_count}")
        except Exception as e:
            print(f"百炼平台文件清理失败: {e}")
    except Exception as e:
        print(f"上传或分析失败: {e}")

def main():
    print("欢迎使用CVE数据抓取和处理工具")
    print("=" * 50)
    start_date, end_date = get_date_input()
    print(f"\n已选择日期范围: {start_date} 至 {end_date}")
    try:
        with open('components.txt', 'r', encoding='utf-8') as f:
            components = [line.strip() for line in f if line.strip() and not line.startswith('#')]
    except Exception as e:
        print(f"读取组件文件出错: {e}")
        return
    if not components:
        print("没有找到任何组件名称")
        return
    print(f"\n找到 {len(components)} 个组件")
    print("组件列表:", components)
    if input("\n是否开始抓取数据？(y/n): ").lower() != 'y':
        print("操作已取消")
        return
    all_results = []
    with ThreadPoolExecutor(max_workers=8) as pool:
        futures = [pool.submit(fetch_cve_data, c, start_date, end_date) for c in components]
        for fut in as_completed(futures):
            all_results.append(fut.result())
    print("\n开始处理数据并生成Excel文件...")
    xlsx_path = write_excel(all_results)
    return xlsx_path

if __name__ == "__main__":
    xlsx_path = main()
    if xlsx_path and input("\n是否上传最新Excel到百炼平台并用大模型分析？(y/n): ").lower() == 'y':
        try:
            prompt = read_prompt()
        except Exception as e:
            print(str(e))
        else:
            upload_and_analyze_bailian(xlsx_path, prompt) 